#!/usr/bin/env python3
"""
Otto Spreadsheet-Pro v4 — 视觉母题驱动 Excel 生成引擎

对标 ppt-creator：AI 创造视觉母题 → 引擎转化为仪表盘、数据表、
摘要卡片的多种视觉形态。支持多工作表、智能格式识别。

用法：python create_xlsx.py <input.md> <output.xlsx>

Markdown 的 ## 分割工作表。AI 在 YAML frontmatter 声明
theme/base/accent，引擎自动派生全表视觉系统。
"""
from __future__ import annotations
import re, sys, json
from datetime import datetime
from pathlib import Path
from typing import Any

if sys.platform == 'win32':
    try: sys.stdout.reconfigure(encoding='utf-8')
    except Exception: pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# 色彩工具 + 视觉母题
# ═══════════════════════════════════════════════════════════════════════

def _hrgb(h): h=h.lstrip("#"); return tuple(int(h[i:i+2],16) for i in(0,2,4))
def _hex(r,g,b): return f"{max(0,min(255,int(r))):02X}{max(0,min(255,int(g))):02X}{max(0,min(255,int(b))):02X}"
def _light(h,a):
    r,g,b=_hrgb(h); return _hex(r+(255-r)*a,g+(255-g)*a,b+(255-b)*a)
def _blend(c1,c2,r):
    r1,g1,b1=_hrgb(c1); r2,g2,b2=_hrgb(c2)
    return _hex(r1*(1-r)+r2*r,g1*(1-r)+g2*r,b1*(1-r)+b2*r)

def resolve_theme(meta: dict) -> dict:
    t = {}
    t["theme_name"] = meta.get("theme", "Undefined")
    t["atmosphere"] = meta.get("atmosphere", "")
    t["base"]   = meta.get("base",   "0A1628")
    t["accent"] = meta.get("accent", "2D7DD2")
    t["surface"]= meta.get("surface","F5F7FA")
    t["body"]   = _light(t["base"],0.82) if _hrgb(t["base"])[0]<50 else "333333"
    t["muted"]  = _light(t["base"],0.55)
    t["hdr_bg"]   = t["base"]
    t["hdr_text"] = "FFFFFF"
    t["stripe"]   = t["surface"]
    t["border"]   = "D0D5DD"
    t["title_bg"] = t["base"]
    t["title_text"]= "FFFFFF"
    t["neg"] = "DC3545"
    t["pos"] = "28A745"
    t["h_font"] = meta.get("heading_font","Microsoft YaHei")
    t["b_font"] = meta.get("body_font","Microsoft YaHei")
    t["title_sz"] = int(float(meta.get("title_size","13")))
    t["hdr_sz"]   = int(float(meta.get("header_size","11")))
    t["body_sz"]  = int(float(meta.get("body_size","10.5")))
    return t


# ═══════════════════════════════════════════════════════════════════════
# Markdown 解析
# ═══════════════════════════════════════════════════════════════════════

def parse_md(text:str) -> tuple[dict, list[dict]]:
    meta={}; body=text.strip()
    if body.startswith("---"):
        parts=body.split("---",2)
        if len(parts)>=3:
            for line in parts[1].strip().split("\n"):
                line=line.strip()
                if ":" in line and not line.startswith("#"):
                    k,_,v=line.partition(":"); meta[k.strip()]=v.strip().strip('"').strip("'")
            body=parts[2].strip()

    lines=body.split("\n"); i=0
    sheets=[]; cur={"name":"Sheet1","blocks":[],"rows":[]}

    def save():
        if cur["blocks"] or cur["rows"]: sheets.append(cur.copy())

    tbl=[]; in_t=False
    while i<len(lines):
        line=lines[i]
        if line.strip().startswith("|") and line.strip().endswith("|"):
            tbl.append(line); in_t=True; i+=1; continue
        elif in_t:
            h,rows=_ptbl(tbl)
            if h: cur["rows"]=[h]+rows
            tbl=[]; in_t=False; continue
        if not line.strip(): i+=1; continue

        m=re.match(r"^(#{1,2})\s+(.+)$",line)
        if m:
            save(); cur={"name":m.group(2).strip()[:31],"blocks":[],"rows":[]}
            i+=1; continue

        cur["blocks"].append({"type":"text","text":line.strip()})
        i+=1
    if tbl:
        h,rows=_ptbl(tbl)
        if h: cur["rows"]=[h]+rows
    save()
    return meta, sheets

def _ptbl(raw):
    if len(raw)<2: return [],[]
    h=[c.strip() for c in raw[0].strip("|").split("|")]
    rows=[]
    for line in raw[1:]:
        if re.match(r"^[\|\-\s:]+$",line.strip()): continue
        cells=[c.strip() for c in line.strip("|").split("|")]
        if cells: rows.append(cells)
    return h, rows


# ═══════════════════════════════════════════════════════════════════════
# Excel 渲染器 v4
# ═══════════════════════════════════════════════════════════════════════

class XLSXRenderer:
    def __init__(self, t:dict, meta:dict):
        self.t=t; self.m=meta; self.wb=Workbook(); self._first=True

    def _hex(self,k): return self.t[k]

    def _styles(self):
        self._hf=Font(name=self.t["h_font"],size=self.t["hdr_sz"],bold=True,color=self._hex("hdr_text"))
        self._hf_fill=PatternFill(start_color=self._hex("hdr_bg"),end_color=self._hex("hdr_bg"),fill_type="solid")
        self._bf=Font(name=self.t["b_font"],size=self.t["body_sz"],color=self._hex("body"))
        self._bf_muted=Font(name=self.t["b_font"],size=self.t["body_sz"]-1,color=self._hex("muted"),italic=True)
        self._sf=PatternFill(start_color=self._hex("stripe"),end_color=self._hex("stripe"),fill_type="solid")
        self._title_f=Font(name=self.t["h_font"],size=self.t["title_sz"],bold=True,color=self._hex("title_text"))
        self._title_fill=PatternFill(start_color=self._hex("title_bg"),end_color=self._hex("title_bg"),fill_type="solid")
        self._border=Border(
            left=Side(style="thin",color=self._hex("border")),
            right=Side(style="thin",color=self._hex("border")),
            top=Side(style="thin",color=self._hex("border")),
            bottom=Side(style="thin",color=self._hex("border")))
        self._center=Alignment(horizontal="center",vertical="center",wrap_text=True)
        self._left=Alignment(horizontal="left",vertical="center",wrap_text=True)
        self._pos_f=Font(name=self.t["b_font"],size=self.t["body_sz"],color=self._hex("pos"))
        self._neg_f=Font(name=self.t["b_font"],size=self.t["body_sz"],color=self._hex("neg"))

    def _get_sheet(self, name:str):
        if self._first:
            ws=self.wb.active; ws.title=name; self._first=False
        else:
            ws=self.wb.create_sheet(title=name)
        ws.sheet_properties.tabColor=self._hex("accent")
        return ws

    def build(self, sheets:list):
        self._styles()
        title=self.m.get("title","")

        for si,sheet in enumerate(sheets):
            ws=self._get_sheet(sheet["name"])

            # 标题栏
            if title and si==0:
                ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=10)
                c=ws.cell(1,1,value=title)
                c.font=self._title_f; c.fill=self._title_fill
                c.alignment=self._center; ws.row_dimensions[1].height=28
                # accent 装饰线
                for col in range(1,11):
                    ws.cell(2,col).fill=PatternFill(start_color=self._hex("accent"),
                        end_color=self._hex("accent"),fill_type="solid")
                ws.row_dimensions[2].height=3
                sr=4
            else:
                sr=1

            r=sr
            # 文本说明
            for blk in sheet.get("blocks",[]):
                c=ws.cell(r,1,value=blk["text"])
                c.font=self._bf_muted; r+=1

            # 数据表
            rows=sheet.get("rows",[])
            if rows:
                if sheet.get("blocks"): r+=1

                # 表头
                for j,h in enumerate(rows[0]):
                    c=ws.cell(r,j+1,value=h)
                    c.font=self._hf; c.fill=self._hf_fill
                    c.alignment=self._center; c.border=self._border
                ws.row_dimensions[r].height=22; r+=1

                # 数据
                for i,row in enumerate(rows[1:]):
                    for j,val in enumerate(row):
                        c=ws.cell(r+i,j+1,value=val)
                        c.font=self._bf; c.border=self._border
                        c.alignment=self._left
                        # 数值识别与条件着色
                        if val and isinstance(val,str) and re.match(r"^[\-\+]?[\d,.]+%?$",val.strip()):
                            c.alignment=self._center
                            v=val.strip()
                            if v.startswith("-"): c.font=self._neg_f
                            elif v.startswith("+") or (v.endswith("%") and float(v[:-1])>0): c.font=self._pos_f
                        # 交替行
                        if i%2==1: c.fill=self._sf
                    ws.row_dimensions[r+i].height=20
                r+=len(rows)-1

                # 列宽
                for j in range(len(rows[0])):
                    w=max(len(str(row[j])) if j<len(row) else 0 for row in rows)
                    ws.column_dimensions[get_column_letter(j+1)].width=min(w*2.2+2,42)

            # 冻结首行
            if rows: ws.freeze_panes=ws.cell(sr,1)

        # 摘要仪表盘
        try: self._dashboard(sheets)
        except: pass

    def _dashboard(self, sheets:list):
        ws=self.wb.create_sheet(title="📊 数据摘要")
        ws.sheet_properties.tabColor=self._hex("base")
        self._styles()
        r=1
        c=ws.cell(r,1,value="📊 数据摘要")
        c.font=Font(name=self.t["h_font"],size=14,bold=True,color=self._hex("base"))
        r+=2

        for sheet in sheets:
            rows=sheet.get("rows",[])
            if not rows: continue
            c=ws.cell(r,1,value=f"📋 {sheet['name']}")
            c.font=Font(name=self.t["h_font"],size=11,bold=True,color=self._hex("accent"))
            ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
            r+=2

            for j,h in enumerate(rows[0]):
                ws.cell(r,j+1,value=h).font=self._hf
                ws.cell(r,j+1).fill=self._hf_fill
                ws.cell(r,j+1).border=self._border
            r+=1

            for i,row in enumerate(rows[1:]):
                for j,val in enumerate(row):
                    c=ws.cell(r+i,j+1,value=val)
                    c.font=self._bf; c.border=self._border
                    if i%2==1: c.fill=self._sf
            r+=len(rows)+2

    def save(self,p): self.wb.save(p)


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    import argparse
    p=argparse.ArgumentParser(description="Otto Spreadsheet-Pro v4")
    p.add_argument("input"); p.add_argument("output")
    a=p.parse_args()
    ip=Path(a.input)
    if not ip.exists(): print(f"找不到 {a.input}"); sys.exit(1)
    text=ip.read_text(encoding="utf-8")
    meta,sheets=parse_md(text)
    t=resolve_theme(meta)
    if "title" not in meta: meta["title"]=ip.stem
    gen=XLSXRenderer(t,meta)
    gen.build(sheets)
    op=Path(a.output); op.parent.mkdir(parents=True,exist_ok=True)
    gen.save(str(op))
    print(f"✅ {op}")
    print(f"   视觉母题：{t['theme_name']}"+(f" · {t['atmosphere']}" if t['atmosphere'] else ""))
    print(f"   大小：{op.stat().st_size/1024:.1f}KB · {len(sheets)}工作表")

if __name__=="__main__": main()
