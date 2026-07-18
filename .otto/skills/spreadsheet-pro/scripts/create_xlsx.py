#!/usr/bin/env python3
"""
Otto Spreadsheet-Pro v3：AI 驱动 Excel 生成引擎

设计令牌 → 专业 .xlsx。AI 创造表格视觉语言，引擎渲染。
基于 openpyxl。

用法：
  python create_xlsx.py <input.md> <output.xlsx>

输入 Markdown 的 YAML frontmatter 指定设计令牌。
支持多 sheet（用 ## 级标题分割）。
"""
from __future__ import annotations

import json, os, re, sys
from datetime import datetime
from pathlib import Path
from typing import Any

if sys.platform == 'win32':
    try: sys.stdout.reconfigure(encoding='utf-8')
    except Exception: pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
except ImportError:
    print("错误：需要 openpyxl。pip install openpyxl"); sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# 设计令牌
# ═══════════════════════════════════════════════════════════════════════

def parse_tokens(meta: dict) -> dict:
    t: dict[str, Any] = {}
    t["design_name"] = meta.get("design_name", "")
    t["design_mood"] = meta.get("design_mood", "")

    t["primary"]          = meta.get("primary",           "1B3A5C")
    t["accent"]           = meta.get("accent",            "2E75B6")
    t["body_color"]       = meta.get("body_color",        "333333")
    t["muted"]            = meta.get("muted",             "888888")
    t["header_bg"]        = meta.get("header_bg",         t["primary"])
    t["header_text"]      = meta.get("header_text",       "FFFFFF")
    t["stripe_bg"]        = meta.get("stripe_bg",         "F0F4F8")
    t["border_color"]     = meta.get("border_color",      "CCCCCC")
    t["title_bg"]         = meta.get("title_bg",          t["primary"])
    t["title_text"]       = meta.get("title_text",        "FFFFFF")
    t["summary_bg"]       = meta.get("summary_bg",        "F8F9FA")
    t["negative_color"]   = meta.get("negative_color",    "DC3545")
    t["positive_color"]   = meta.get("positive_color",    "28A745")

    t["heading_font"] = meta.get("heading_font", "Microsoft YaHei")
    t["body_font"]    = meta.get("body_font",    "Microsoft YaHei")
    t["title_size"]   = int(float(meta.get("title_size", "14")))
    t["header_size"]  = int(float(meta.get("header_size", "11")))
    t["body_size"]    = int(float(meta.get("body_size", "10.5")))

    t["sheet_names"] = meta.get("sheet_names", "").split(",") if meta.get("sheet_names") else []

    return t


def _hex(h: str) -> str:
    return h.lstrip("#")


# ═══════════════════════════════════════════════════════════════════════
# Markdown 解析
# ═══════════════════════════════════════════════════════════════════════

def parse_frontmatter(text: str) -> tuple[dict, str]:
    meta = {}; body = text.strip()
    if body.startswith("---"):
        parts = body.split("---", 2)
        if len(parts) >= 3:
            for line in parts[1].strip().split("\n"):
                line = line.strip()
                if ":" in line and not line.startswith("#"):
                    k, _, v = line.partition(":")
                    meta[k.strip()] = v.strip().strip('"').strip("'")
            body = parts[2].strip()
    return meta, body


def parse_body(text: str) -> list[dict]:
    """解析为 sheet 级 + 表格/段落。"""
    sections: list[dict] = []
    lines = text.split("\n"); i = 0
    current_sheet = {"name": "Sheet1", "blocks": [], "rows": [], "title_row": None}

    def save_section():
        if current_sheet["blocks"] or current_sheet["rows"]:
            sections.append(current_sheet.copy())

    while i < len(lines):
        line = lines[i]

        # H1/H2 作为新 sheet
        m = re.match(r"^(#{1,2})\s+(.+)$", line)
        if m:
            save_section()
            current_sheet = {"name": m.group(2).strip()[:31], "blocks": [], "rows": []}
            i += 1; continue

        # 表格
        if line.strip().startswith("|") and line.strip().endswith("|"):
            tbl = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                tbl.append(lines[i].strip())
                i += 1
            hdrs, rows = _parse_table(tbl)
            if hdrs:
                current_sheet["rows"] = [hdrs] + rows
            continue

        if not line.strip(): i += 1; continue

        # 普通段落
        current_sheet["blocks"].append({"type": "text", "text": line.strip()})
        i += 1

    save_section()
    return sections


def _parse_table(raw: list[str]) -> tuple[list[str], list[list[str]]]:
    if len(raw) < 2: return [], []
    hdrs = [c.strip() for c in raw[0].strip("|").split("|")]
    rows = []
    for r in raw[1:]:
        if re.match(r"^[\|\-\s:]+$", r): continue
        cells = [c.strip() for c in r.strip("|").split("|")]
        if cells: rows.append(cells)
    return hdrs, rows


# ═══════════════════════════════════════════════════════════════════════
# Excel 渲染器
# ═══════════════════════════════════════════════════════════════════════

class XLSXRenderer:
    def __init__(self, tokens: dict, meta: dict):
        self.t = tokens; self.m = meta
        self.wb = Workbook()
        self._first = True

    def _styles(self):
        """返回常用样式。"""
        hdr_font = Font(name=self.t["heading_font"], size=self.t["header_size"],
                         bold=True, color=_hex(self.t["header_text"]))
        hdr_fill = PatternFill(start_color=_hex(self.t["header_bg"]),
                                end_color=_hex(self.t["header_bg"]), fill_type="solid")
        body_font = Font(name=self.t["body_font"], size=self.t["body_size"],
                          color=_hex(self.t["body_color"]))
        stripe_fill = PatternFill(start_color=_hex(self.t["stripe_bg"]),
                                   end_color=_hex(self.t["stripe_bg"]), fill_type="solid")
        title_font = Font(name=self.t["heading_font"], size=self.t["title_size"],
                           bold=True, color=_hex(self.t["title_text"]))
        title_fill = PatternFill(start_color=_hex(self.t["title_bg"]),
                                  end_color=_hex(self.t["title_bg"]), fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color=_hex(self.t["border_color"])),
            right=Side(style="thin", color=_hex(self.t["border_color"])),
            top=Side(style="thin", color=_hex(self.t["border_color"])),
            bottom=Side(style="thin", color=_hex(self.t["border_color"])),
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self._hdr_font = hdr_font; self._hdr_fill = hdr_fill
        self._body_font = body_font; self._stripe_fill = stripe_fill
        self._title_font = title_font; self._title_fill = title_fill
        self._border = thin_border
        self._center = center; self._left = left_align

    def _get_sheet(self, name: str):
        if self._first:
            ws = self.wb.active
            ws.title = name
            self._first = False
        else:
            ws = self.wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = _hex(self.t["accent"])
        return ws

    def build(self, sections: list[dict]):
        self._styles()
        title = self.m.get("title", "")

        for sec in sections:
            ws = self._get_sheet(sec["name"])

            # 标题行
            if title and ws.title == self.wb.sheetnames[0]:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
                c = ws.cell(row=1, column=1, value=title)
                c.font = self._title_font; c.fill = self._title_fill
                c.alignment = self._center
                ws.row_dimensions[1].height = 30
                start_row = 3
            else:
                start_row = 1

            # 文本块
            r = start_row
            for blk in sec.get("blocks", []):
                c = ws.cell(row=r, column=1, value=blk["text"])
                c.font = Font(name=self.t["body_font"], size=self.t["body_size"],
                              color=_hex(self.t["muted"]), italic=True)
                r += 1

            # 表格
            rows = sec.get("rows", [])
            if rows:
                if sec.get("blocks"): r += 1  # 分隔
                # 表头
                for j, h in enumerate(rows[0]):
                    c = ws.cell(row=r, column=j + 1, value=h)
                    c.font = self._hdr_font; c.fill = self._hdr_fill
                    c.alignment = self._center; c.border = self._border
                ws.row_dimensions[r].height = 22
                r += 1

                # 数据行
                for i, row in enumerate(rows[1:]):
                    for j, val in enumerate(row):
                        c = ws.cell(row=r + i, column=j + 1, value=val)
                        c.font = self._body_font
                        c.border = self._border
                        c.alignment = self._left
                        # 尝试识别数字
                        if val and re.match(r"^-?[\d,.]+%?$", val.strip()):
                            c.alignment = self._center
                            if val.strip().startswith("-"):
                                c.font = Font(name=self.t["body_font"],
                                              size=self.t["body_size"],
                                              color=_hex(self.t["negative_color"]))
                        # 条纹
                        if i % 2 == 1:
                            c.fill = self._stripe_fill
                    ws.row_dimensions[r + i].height = 20

                r += len(rows) - 1

            # 列宽自适应
            if rows:
                for j in range(len(rows[0])):
                    max_len = max(len(str(row[j])) if j < len(row) else 0 for row in rows)
                    ws.column_dimensions[get_column_letter(j + 1)].width = min(max_len * 1.8 + 2, 40)

        # 添加摘要 sheet
        if title:
            try:
                self._add_summary(sections)
            except Exception:
                pass

    def _add_summary(self, sections: list[dict]):
        ws = self.wb.create_sheet(title="📊 数据摘要")
        ws.sheet_properties.tabColor = _hex(self.t["primary"])
        r = 1
        c = ws.cell(row=r, column=1, value="📊 数据摘要")
        c.font = Font(name=self.t["heading_font"], size=14, bold=True,
                      color=_hex(self.t["primary"]))
        r += 2
        for sec in sections:
            rows = sec.get("rows", [])
            if not rows: continue
            c = ws.cell(row=r, column=1, value=f"📋 {sec['name']}")
            c.font = Font(name=self.t["heading_font"], size=11, bold=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            r += 2

            for j, h in enumerate(rows[0]):
                ws.cell(row=r, column=j+1, value=h).font = self._hdr_font
                ws.cell(row=r, column=j+1).fill = self._hdr_fill
                ws.cell(row=r, column=j+1).border = self._border
            r += 1

            for i, row in enumerate(rows[1:]):
                for j, val in enumerate(row):
                    ws.cell(row=r+i, column=j+1, value=val).font = self._body_font
                    ws.cell(row=r+i, column=j+1).border = self._border
            r += len(rows) + 1

    def save(self, path: str):
        self.wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

def main():
    import argparse
    p = argparse.ArgumentParser(description="Otto Spreadsheet-Pro: AI 驱动 Excel 生成引擎")
    p.add_argument("input", help="输入 Markdown 文件")
    p.add_argument("output", help="输出 .xlsx 文件")
    a = p.parse_args()

    in_path = Path(a.input)
    if not in_path.exists():
        print(f"错误：找不到 {a.input}"); sys.exit(1)

    text = in_path.read_text(encoding="utf-8")
    meta, body = parse_frontmatter(text)
    tokens = parse_tokens(meta)
    if "title" not in meta: meta["title"] = in_path.stem

    sections = parse_body(body)
    gen = XLSXRenderer(tokens, meta)
    gen.build(sections)

    out = Path(a.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    gen.save(str(out))

    dn = tokens.get("design_name",""); dm = tokens.get("design_mood","")
    tag = ""
    if dn: tag += dn
    if dm: tag += (" · " + dm) if tag else dm
    print(f"✅ Excel 已生成：{out}")
    print(f"   视觉语言：{tag}" if tag else "   视觉语言：自动推断")
    print(f"   大小：{out.stat().st_size / 1024:.1f} KB · {len(sections)} 个工作表")


if __name__ == "__main__":
    main()
